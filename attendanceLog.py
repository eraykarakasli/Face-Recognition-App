
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from datetime import date


today = date.today()
day = today.strftime("%b-%d-%Y")

day_str = "attendance-" + day + ".csv"
print(day_str)
data = pd.read_csv('day_str')

wb = openpyxl.Workbook()
sayfa = wb.active

a2 = len(data)           ### toplan satır sayısı
a3 = len(data.columns)   ### toplam sütun sayısı
print('satır uzunluğu: ', a2)
print('sütun sayısı: ', a3)


for x in range(a3):      ### sütun başlıklarını yazdırma döngüsü
    c = x + 1
    sayfa.cell(row = 1, column = c).value = data.columns[x]


for x in range(a2):    ### tüm satırlardaki verileri excele yazdırma döngüsü
    for y in range(a3):
        r = x + 2
        c = y + 1
        sayfa.cell(row = r, column = c).value = data.iat[x,y]


wb.save("yeni_excel_dosyasi.xlsx")

print('İşlem başarıyla tamamlandı. Excel dosyanız oluşturuldu')